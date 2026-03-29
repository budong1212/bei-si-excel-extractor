#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
贝思excel内容提取工具
支持多文件、多表格、百万行大文件、拖拽、实时进度、智能表头合并、流式内存优化
"""

import os
import sys
import time
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    HAS_DND = True
except ImportError:
    HAS_DND = False


class ExcelExtractorApp:
    def __init__(self):
        self.root = TkinterDnD.Tk() if HAS_DND else tk.Tk()
        self.root.title("贝思excel内容提取工具")
        self.root.geometry("900x700")
        self.root.minsize(800, 600)
        self.root.configure(bg="#f0f4f8")
        self.file_list = []
        self.is_running = False
        self.cancel_flag = False
        self._build_ui()

    # ================================================================== UI ==
    def _build_ui(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TButton", font=("微软雅黑", 10), padding=6)
        style.configure("TLabel", background="#f0f4f8", font=("微软雅黑", 10))
        style.configure("TLabelframe", background="#f0f4f8")
        style.configure("TLabelframe.Label", background="#f0f4f8", font=("微软雅黑", 10, "bold"))
        style.configure("green.Horizontal.TProgressbar", troughcolor="#dce3eb", background="#4caf50")

        main = tk.Frame(self.root, bg="#f0f4f8", padx=12, pady=12)
        main.pack(fill=tk.BOTH, expand=True)

        # 文件列表
        ff = ttk.LabelFrame(main, text=" 📂 文件列表 ", padding=8)
        ff.pack(fill=tk.BOTH, expand=True)
        btn_bar = tk.Frame(ff, bg="#f0f4f8")
        btn_bar.pack(fill=tk.X, pady=(0, 6))
        ttk.Button(btn_bar, text="➕ 添加文件",   command=self._add_files).pack(side=tk.LEFT, padx=(0,6))
        ttk.Button(btn_bar, text="📁 添加文件夹", command=self._add_folder).pack(side=tk.LEFT, padx=(0,6))
        ttk.Button(btn_bar, text="🗑 清空列表",  command=self._clear_files).pack(side=tk.LEFT)
        lc = tk.Frame(ff, bg="#f0f4f8")
        lc.pack(fill=tk.BOTH, expand=True)
        self.file_listbox = tk.Listbox(
            lc, selectmode=tk.EXTENDED, bg="#fff", fg="#333",
            font=("微软雅黑", 9), relief=tk.FLAT,
            highlightthickness=1, highlightcolor="#4caf50",
            highlightbackground="#ccd", height=8)
        sbx = ttk.Scrollbar(lc, orient=tk.HORIZONTAL, command=self.file_listbox.xview)
        sby = ttk.Scrollbar(lc, orient=tk.VERTICAL,   command=self.file_listbox.yview)
        self.file_listbox.configure(xscrollcommand=sbx.set, yscrollcommand=sby.set)
        sby.pack(side=tk.RIGHT, fill=tk.Y)
        sbx.pack(side=tk.BOTTOM, fill=tk.X)
        self.file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        hint = "(支持拖拽文件/文件夹到此区域)" if HAS_DND else "(安装 tkinterdnd2 后可拖拽)"
        tk.Label(ff, text=hint, bg="#f0f4f8", fg="#888", font=("微软雅黑", 8)).pack(anchor=tk.W)
        if HAS_DND:
            self.file_listbox.drop_target_register(DND_FILES)
            self.file_listbox.dnd_bind("<<Drop>>", self._on_drop)

        # 关键词
        kf = ttk.LabelFrame(main, text=" 🔍 搜索关键词（每行一个） ", padding=8)
        kf.pack(fill=tk.X, pady=(10, 0))
        self.keyword_text = tk.Text(
            kf, height=5, font=("微软雅黑", 10),
            bg="#fff", relief=tk.FLAT,
            highlightthickness=1, highlightbackground="#ccd")
        ks = ttk.Scrollbar(kf, orient=tk.VERTICAL, command=self.keyword_text.yview)
        self.keyword_text.configure(yscrollcommand=ks.set)
        ks.pack(side=tk.RIGHT, fill=tk.Y)
        self.keyword_text.pack(fill=tk.X)

        # 匹配方式
        of = tk.Frame(main, bg="#f0f4f8")
        of.pack(fill=tk.X, pady=(8, 0))
        self.match_mode = tk.StringVar(value="contains")
        tk.Label(of, text="匹配方式：", bg="#f0f4f8", font=("微软雅黑", 10)).pack(side=tk.LEFT)
        ttk.Radiobutton(of, text="包含",     variable=self.match_mode, value="contains").pack(side=tk.LEFT, padx=4)
        ttk.Radiobutton(of, text="完全匹配", variable=self.match_mode, value="exact").pack(side=tk.LEFT, padx=4)
        ttk.Radiobutton(of, text="开头匹配", variable=self.match_mode, value="startswith").pack(side=tk.LEFT, padx=4)
        self.case_sensitive = tk.BooleanVar(value=False)
        ttk.Checkbutton(of, text="区分大小写", variable=self.case_sensitive).pack(side=tk.LEFT, padx=12)
        self.whole_word = tk.BooleanVar(value=False)
        ttk.Checkbutton(of, text="全词匹配（带分隔符）", variable=self.whole_word).pack(side=tk.LEFT, padx=4)

        # 输出路径
        outf = tk.Frame(main, bg="#f0f4f8")
        outf.pack(fill=tk.X, pady=(8, 0))
        tk.Label(outf, text="输出文件：", bg="#f0f4f8", font=("微软雅黑", 10)).pack(side=tk.LEFT)
        self.output_var = tk.StringVar(value=str(Path.home() / "贝思提取结果.xlsx"))
        tk.Entry(outf, textvariable=self.output_var, font=("微软雅黑", 9), width=50).pack(side=tk.LEFT, padx=6)
        ttk.Button(outf, text="浏览", command=self._choose_output).pack(side=tk.LEFT)

        # 进度
        pf = ttk.LabelFrame(main, text=" ⚡ 进度 ", padding=8)
        pf.pack(fill=tk.X, pady=(10, 0))
        self.progress_bar = ttk.Progressbar(
            pf, style="green.Horizontal.TProgressbar",
            orient=tk.HORIZONTAL, length=100, mode="determinate")
        self.progress_bar.pack(fill=tk.X)
        ir = tk.Frame(pf, bg="#f0f4f8")
        ir.pack(fill=tk.X, pady=(4, 0))
        self.progress_label = tk.Label(ir, text="就绪", bg="#f0f4f8", fg="#555", font=("微软雅黑", 9))
        self.progress_label.pack(side=tk.LEFT)
        self.speed_label = tk.Label(ir, text="", bg="#f0f4f8", fg="#4caf50", font=("微软雅黑", 9, "bold"))
        self.speed_label.pack(side=tk.RIGHT)

        # 按钮
        af = tk.Frame(main, bg="#f0f4f8")
        af.pack(pady=(12, 0))
        self.start_btn  = ttk.Button(af, text="▶  开始提取", command=self._start_extract)
        self.start_btn.pack(side=tk.LEFT, padx=8)
        self.cancel_btn = ttk.Button(af, text="⏹  取消", command=self._cancel, state=tk.DISABLED)
        self.cancel_btn.pack(side=tk.LEFT, padx=8)

    # ============================================================= 文件管理 ==
    def _add_files(self):
        paths = filedialog.askopenfilenames(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls *.xlsm"), ("所有文件", "*.*")])
        for p in paths:
            self._add_path(p)

    def _add_folder(self):
        folder = filedialog.askdirectory(title="选择文件夹")
        if folder:
            for rd, _, files in os.walk(folder):
                for f in files:
                    if f.lower().endswith((".xlsx", ".xls", ".xlsm")):
                        self._add_path(os.path.join(rd, f))

    def _add_path(self, path):
        if path not in self.file_list:
            self.file_list.append(path)
            self.file_listbox.insert(tk.END, path)

    def _clear_files(self):
        self.file_list.clear()
        self.file_listbox.delete(0, tk.END)

    def _on_drop(self, event):
        for p in self.root.tk.splitlist(event.data):
            p = p.strip()
            if os.path.isdir(p):
                for rd, _, files in os.walk(p):
                    for f in files:
                        if f.lower().endswith((".xlsx", ".xls", ".xlsm")):
                            self._add_path(os.path.join(rd, f))
            elif os.path.isfile(p) and p.lower().endswith((".xlsx", ".xls", ".xlsm")):
                self._add_path(p)

    def _choose_output(self):
        p = filedialog.asksaveasfilename(
            title="保存结果",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")])
        if p:
            self.output_var.set(p)

    # ============================================================= 匹配逻辑 ==
    def _row_matches(self, row_values, keywords, mode, case, whole_word):
        cells = [str(v) if v is not None else "" for v in row_values]
        for kw in keywords:
            kw_cmp = kw if case else kw.lower()
            for v in cells:
                v_cmp = v if case else v.lower()
                if mode == "contains":
                    if whole_word:
                        import re
                        if re.search(r'(?<![\w一-鿿])' + re.escape(kw_cmp) + r'(?![\w一-鿿])', v_cmp):
                            return True
                    else:
                        if kw_cmp in v_cmp:
                            return True
                elif mode == "exact":
                    if v_cmp.strip() == kw_cmp:
                        return True
                elif mode == "startswith":
                    if v_cmp.startswith(kw_cmp):
                        return True
        return False

    # ============================================================= 启动 ======
    def _start_extract(self):
        if not self.file_list:
            messagebox.showwarning("提示", "请先添加Excel文件！"); return
        kws = [l.strip() for l in self.keyword_text.get("1.0", tk.END).splitlines() if l.strip()]
        if not kws:
            messagebox.showwarning("提示", "请输入关键词！"); return
        out = self.output_var.get().strip()
        if not out:
            messagebox.showwarning("提示", "请设置输出文件！"); return

        self.is_running = True
        self.cancel_flag = False
        self.start_btn.config(state=tk.DISABLED)
        self.cancel_btn.config(state=tk.NORMAL)
        self.progress_bar["value"] = 0

        threading.Thread(
            target=self._extract_worker,
            args=(list(self.file_list), kws, out,
                  self.match_mode.get(),
                  self.case_sensitive.get(),
                  self.whole_word.get()),
            daemon=True).start()

    def _cancel(self):
        self.cancel_flag = True
        self._update_status("正在取消...")

    def _update_status(self, msg, progress=None, speed=None):
        def _do():
            self.progress_label.config(text=msg)
            if progress is not None:
                self.progress_bar["value"] = progress
            if speed is not None:
                self.speed_label.config(text=speed)
        self.root.after(0, _do)

    def _finish(self, success, output_path=None):
        def _do():
            self.is_running = False
            self.start_btn.config(state=tk.NORMAL)
            self.cancel_btn.config(state=tk.DISABLED)
            self.speed_label.config(text="")
            if success:
                self.progress_bar["value"] = 100
                if messagebox.askyesno("完成",
                    f"提取完成！\n结果已保存至：\n{output_path}\n\n是否立即打开？"):
                    if sys.platform == "win32":
                        os.startfile(output_path)
                    else:
                        os.system(f'open "{output_path}"')
            else:
                self.progress_bar["value"] = 0
        self.root.after(0, _do)

    # ============================================================= 工作线程 ==
    def _extract_worker(self, files, keywords, output_path, mode, case, whole_word):
        """
        两阶段流式处理，适合百万行大文件：
          阶段一：扫描所有文件，收集「表头分组顺序」并将命中行暂存到临时文件
          阶段二：按表头分组顺序将临时文件内容写入输出 Excel
        内存常驻量不随文件大小增长。
        """
        import tempfile, csv, json

        total_files = len(files)
        start_time  = time.time()
        total_scanned = 0
        total_matched = 0

        # 临时目录：每个 header_key 对应一个 csv临时文件
        tmpdir = tempfile.mkdtemp(prefix="beisi_extract_")
        # header_order: [(header_tuple, tmp_csv_path), ...]
        header_order = []
        header_index  = {}   # header_tuple -> tmp_csv_path
        header_writer = {}   # header_tuple -> csv.writer
        header_fh     = {}   # header_tuple -> file handle

        def _get_writer(header):
            if header not in header_index:
                tmp_path = os.path.join(tmpdir, f"group_{len(header_index)}.csv")
                fh = open(tmp_path, "w", newline="", encoding="utf-8")
                writer = csv.writer(fh)
                header_index[header]  = tmp_path
                header_writer[header] = writer
                header_fh[header]     = fh
                header_order.append(header)
            return header_writer[header]

        try:
            # ──────────────── 阶段一：扫描 ────────────────
            for file_idx, filepath in enumerate(files):
                if self.cancel_flag:
                    break
                fname = os.path.basename(filepath)
                self._update_status(
                    f"扫描 [{file_idx+1}/{total_files}]：{fname}",
                    progress=int(file_idx / total_files * 85))

                try:
                    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
                except Exception as e:
                    self._update_status(f"跳过：{fname} — {e}")
                    continue

                for sheet_name in wb.sheetnames:
                    if self.cancel_flag:
                        break
                    ws = wb[sheet_name]
                    rows_iter = ws.iter_rows(values_only=True)
                    try:
                        header_row = next(rows_iter)
                    except StopIteration:
                        continue

                    header = tuple(str(h) if h is not None else "" for h in header_row)
                    writer = _get_writer(header)

                    row_num = 0
                    for row in rows_iter:
                        if self.cancel_flag:
                            break
                        row_num   += 1
                        total_scanned += 1
                        if self._row_matches(row, keywords, mode, case, whole_word):
                            writer.writerow(["" if v is None else v for v in row])
                            total_matched += 1

                        if row_num % 10000 == 0:
                            elapsed = time.time() - start_time
                            spd = total_scanned / elapsed if elapsed else 0
                            self._update_status(
                                f"[{file_idx+1}/{total_files}] {fname}/{sheet_name}  "
                                f"已扫 {total_scanned:,} 行 | 命中 {total_matched:,} 行",
                                progress=int(file_idx / total_files * 85),
                                speed=f"{spd:,.0f} 行/秒")
                wb.close()

            # 关闭所有 csv 句柄
            for fh in header_fh.values():
                fh.close()

            if self.cancel_flag:
                self._update_status("已取消")
                self._finish(False)
                return

            # ──────────────── 阶段二：写入输出 Excel ────────────────
            self._update_status("正在写入输出文件...", progress=88)
            out_wb = Workbook(write_only=True)   # write_only 模式，内存最小
            out_ws = out_wb.create_sheet("提取结果")

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            center = Alignment(horizontal="center", vertical="center")

            def _make_header_cells(hdr):
                cells = []
                for col_name in hdr:
                    c = openpyxl.cell.WriteOnlyCell(out_ws, value=col_name)
                    c.font  = header_font
                    c.fill  = header_fill
                    c.alignment = center
                    cells.append(c)
                return cells

            first = True
            written_rows = 0
            for header in header_order:
                tmp_csv = header_index[header]
                # 检查该分组是否有命中行
                if os.path.getsize(tmp_csv) == 0:
                    continue
                if not first:
                    out_ws.append([None])  # 空行
                first = False

                out_ws.append(_make_header_cells(header))

                with open(tmp_csv, "r", encoding="utf-8", newline="") as f:
                    reader = csv.reader(f)
                    for data_row in reader:
                        out_ws.append(data_row)
                        written_rows += 1
                        if written_rows % 20000 == 0:
                            self._update_status(
                                f"写入中... 已写 {written_rows:,} 行",
                                progress=90 + min(9, written_rows // 50000))

            out_wb.save(output_path)

            elapsed = time.time() - start_time
            spd = total_scanned / elapsed if elapsed else 0
            self._update_status(
                f"完成！扫描 {total_scanned:,} 行 → 提取 {total_matched:,} 行 | 耗时 {elapsed:.1f}s",
                progress=100, speed=f"{spd:,.0f} 行/秒")
            self._finish(True, output_path)

        except Exception as e:
            import traceback
            self._update_status(f"错误：{e}")
            self.root.after(0, lambda: messagebox.showerror(
                "错误", f"提取过程出错：\n{traceback.format_exc()}"))
            self._finish(False)
        finally:
            # 清理临时文件
            import shutil
            try:
                shutil.rmtree(tmpdir, ignore_errors=True)
            except Exception:
                pass

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = ExcelExtractorApp()
    app.run()
